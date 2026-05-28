"""
Scraper for rp5.ru - weekly weather statistics for Central Asia cities.
Uses the "Статистика погоды" tab which provides pre-aggregated
min / max / average temperature for the selected date range.
One request per city, no manual aggregation needed.
"""

import asyncio
import re
import sys
from pathlib import Path

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# English archive URLs - verified working, all have "Weather statistics" tab
# Russian URLs only exist for some cities, English works universally
CITIES = [
    {"country": "Казахстан",   "city": "Актобе",    "url": "https://rp5.ru/Weather_archive_in_Aktobe"},
    {"country": "Казахстан",   "city": "Атырау",    "url": "https://rp5.ru/Weather_archive_in_Atyrau"},
    {"country": "Казахстан",   "city": "Астана",    "url": "https://rp5.ru/Weather_archive_in_Astana"},
    {"country": "Казахстан",   "city": "Караганда", "url": "https://rp5.ru/Weather_archive_in_Karaganda"},
    {"country": "Казахстан",   "city": "Семей",     "url": "https://rp5.ru/Weather_archive_in_Semey"},
    {"country": "Казахстан",   "city": "Шымкент",   "url": "https://rp5.ru/Weather_archive_in_Isfijab_/_Sayram"},
    {"country": "Казахстан",   "city": "Алматы",    "url": "https://rp5.ru/Weather_archive_in_Almaty"},
    {"country": "Узбекистан",  "city": "Нукус",     "url": "https://rp5.ru/Weather_archive_in_Nukus_(airport)"},
    {"country": "Узбекистан",  "city": "Самарканд", "url": "https://rp5.ru/Weather_archive_in_Samarqand_(airport)"},
    {"country": "Узбекистан",  "city": "Ташкент",   "url": "https://rp5.ru/Weather_archive_in_Tashkent_(meteostation)"},
    {"country": "Кыргызстан",  "city": "Бишкек",    "url": "https://rp5.ru/Weather_archive_in_Bishkek"},
    {"country": "Кыргызстан",  "city": "Ош",        "url": "https://rp5.ru/Weather_archive_in_Qorasuv"},
    {"country": "Таджикистан", "city": "Душанбе",   "url": "https://rp5.ru/Weather_archive_in_Dushanbe_(airport)"},
]


def parse_stat_value(text):
    """Extract numeric temperature from stat cell like '-8.6 (16.03.2026)' or '+8.7'."""
    text = text.strip()
    m = re.search(r'([+-]?\d+\.?\d*)', text)
    return float(m.group(1)) if m else None


async def scrape_city(page, city_info, date_from, date_to):
    city = city_info["city"]
    print(f"  {city}...")

    await page.goto(city_info["url"], wait_until='domcontentloaded', timeout=60000)
    await page.wait_for_timeout(2000)

    # Click the statistics tab (English or Russian depending on page language)
    stat_tab = page.locator('text=Weather statistics, text=Статистика погоды').first
    await stat_tab.click()
    await page.wait_for_timeout(1000)

    # Fill date range (StatDate1 = start, StatDate2 = end)
    # These inputs are hidden by CSS - use JavaScript
    await page.evaluate(
        "document.querySelector('input[name=StatDate1]').value = arguments[0]",
        date_from
    )
    await page.evaluate(
        "document.querySelector('input[name=StatDate2]').value = arguments[0]",
        date_to
    )

    # Select "все дни" (s_pe = 1) - all days in range
    await page.evaluate(
        "document.querySelector('input[name=s_pe][value=\"1\"]').click()"
    )
    await page.wait_for_timeout(200)

    # Make sure only T (temperature) parameter is checked (stat_parameter value=1)
    # Uncheck all, then check only T
    await page.evaluate("""
        document.querySelectorAll('input[name=stat_parameter]').forEach(function(cb) {
            cb.checked = (cb.value === '1');
        });
    """)

    # Click "Выполнить расчет" button
    calc_btn = page.locator('text=Выполнить расчет').first
    await calc_btn.click()
    await page.wait_for_timeout(4000)

    # Parse result table
    # Table has columns: Период | Среднее | Минимальное (дата) | Максимальное (дата) | Кол-во набл.
    result = await page.evaluate("""() => {
        // Find the statistics result table - it contains "Среднее значение" or similar header
        var tables = document.querySelectorAll('table');
        for (var t of tables) {
            var html = t.innerHTML;
            if (html.indexOf('реднее') > -1 && html.indexOf('инимальное') > -1) {
                // Get data row (skip header)
                var rows = t.querySelectorAll('tr');
                for (var i = 1; i < rows.length; i++) {
                    var cells = rows[i].querySelectorAll('td');
                    if (cells.length >= 4) {
                        return {
                            avg: cells[1].textContent.trim(),
                            min: cells[2].textContent.trim(),
                            max: cells[3].textContent.trim()
                        };
                    }
                }
            }
        }
        return null;
    }""")

    if not result:
        print(f"    WARN: result table not found for {city}")
        return None

    t_avg = parse_stat_value(result["avg"])
    t_min = parse_stat_value(result["min"])
    t_max = parse_stat_value(result["max"])

    if t_avg is None or t_min is None or t_max is None:
        print(f"    WARN: could not parse values: {result}")
        return None

    print(f"    min={t_min:+.1f}  max={t_max:+.1f}  avg={t_avg:+.1f}")

    return {
        "country": city_info["country"],
        "city": city,
        "t_min": t_min,
        "t_max": t_max,
        "t_avg": t_avg,
    }


def fmt(v):
    return f"+{v:.1f}" if v >= 0 else f"{v:.1f}"


def create_excel(data, date_from, date_to, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Температура"

    hdr_fill = PatternFill("solid", fgColor="4472C4")
    alt1     = PatternFill("solid", fgColor="DCE6F1")
    alt2     = PatternFill("solid", fgColor="B8CCE4")
    hfont    = Font(bold=True, color="FFFFFF", size=10)
    dfont    = Font(size=10)
    bfont    = Font(bold=True, size=11)
    ctr      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    th       = Side(style="thin", color="FFFFFF")
    brd      = Border(left=th, right=th, top=th, bottom=th)

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = (
        f"Температура воздуха в крупных городах средней Азии с\n"
        f"{date_from} по {date_to}, \u00b0C"
    )
    ws["A1"].font = bfont
    ws["A1"].alignment = lft
    ws.row_dimensions[1].height = 42

    # Header row
    for col, h in enumerate(
        ["Страна", "Город", "t мин,\n\u00b0C", "t макс,\n\u00b0C", "t средняя,\n\u00b0C"], 1
    ):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hfont; c.alignment = ctr; c.border = brd
    ws.row_dimensions[2].height = 30

    # Data rows
    for i, row in enumerate(data):
        rn   = i + 3
        fill = alt2 if i % 2 == 0 else alt1
        for col, val in enumerate(
            [row["country"], row["city"], fmt(row["t_min"]), fmt(row["t_max"]), fmt(row["t_avg"])], 1
        ):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill = fill; c.font = dfont; c.alignment = ctr; c.border = brd
        ws.row_dimensions[rn].height = 18

    for col, w in enumerate([16, 14, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(out_path)
    print(f"\nСохранено: {out_path}")


async def main(date_from="09.03.2026", date_to="15.03.2026"):
    out = Path("C:/projects/my-project/ОМТ/Погода/weather_central_asia.xlsx")

    print(f"Сбор статистики погоды {date_from} — {date_to}\n{'='*50}")

    results = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )

        for city_info in CITIES:
            page = await browser.new_page()
            await page.set_extra_http_headers({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            })
            try:
                for attempt in range(3):
                    try:
                        res = await scrape_city(page, city_info, date_from, date_to)
                        if res:
                            results.append(res)
                        break
                    except Exception as e:
                        msg = str(e).split('\n')[0][:70]
                        if attempt < 2:
                            print(f"    повтор ({attempt+1}): {msg}")
                            await page.wait_for_timeout(5000)
                        else:
                            print(f"    ОШИБКА: {msg}")
            finally:
                await page.close()

        await browser.close()

    if results:
        create_excel(results, date_from, date_to, str(out))
        print(f"Готово: {len(results)}/{len(CITIES)} городов")
    else:
        print("Данные не собраны.")


if __name__ == "__main__":
    asyncio.run(main())
