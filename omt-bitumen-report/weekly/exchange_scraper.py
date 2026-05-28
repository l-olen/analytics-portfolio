"""
exchange_scraper.py - Scraper for bitumen exchange data.

Sources:
  UzEx (Uzbekistan): uzex.uz/Trade/List - HTML table, paginated, FULLY WORKING
  ETS (Kazakhstan):  ets.kz/markets/oil_products/trades/ - Playwright (headless Chrome)
  CCX (Kazakhstan):  ccx.kz/gsm - BLOCKED at Kazakhstan national gateway, unavailable

UzEx:
  - Fetches bitumen + mazut records from the last N days
  - Updates sheet "Битум" and "Мазут" in УзБиржа_битум+мазут.xlsx
  - For rows within the last 7 days: updates "Состояние исполнения" (status can change)
  - For rows older: only adds if missing (by deal number)
  - Quantity (col F) is in tons when unit="тонна"; col J/K are Excel-calculated

ETS:
  - Uses Playwright (headless Chromium) -- site requires JavaScript rendering
  - Scrapes weekly aggregate trade data (min/max/avg price, volume, deals)
  - Updates sheet "КАЗ_ETS_База" in ETS_биржа.xlsx
  - NOTE: As of April 2026 bitumen is NOT in ETS oil_products section
    (only diesel, jet fuel visible). ETS data may need manual entry or another source.

Usage:
  python exchange_scraper.py              -- update both
  python exchange_scraper.py --uzex       -- UzEx only
  python exchange_scraper.py --ets        -- ETS only
  python exchange_scraper.py --days 14    -- last N days for UzEx (default 14)
  python exchange_scraper.py --dry-run    -- show fetched data without writing
"""

import argparse
import html
import json
import re
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = Path(__file__).parent
UZEX_FILE = BASE / "УзБиржа_битум+мазут.xlsx"
ETS_FILE = BASE / "ETS_биржа.xlsx"

SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,*/*",
    "Accept-Language": "ru-RU,ru;q=0.9",
}


# ============================================================
# HTTP helper
# ============================================================

def fetch(url: str, timeout: int = 20) -> str:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read().decode("utf-8", errors="replace")


# ============================================================
# UzEx scraper
# ============================================================

UZEX_BASE = "https://uzex.uz/Trade/List"

# Column indices in UzEx HTML table (0-based)
# Headers: Shartnoma sanasi | Bitim raqami | Mahsulot narxi (so'm) | Shartnoma # | Mahsulot markasi | Miqdori | O'lchov birligi | Shartnoma turi | Bitimning bajarilish holati
UZEX_COL_DATE = 0
UZEX_COL_DEAL_NO = 1
UZEX_COL_PRICE = 2
UZEX_COL_CONTRACT_NO = 3
UZEX_COL_PRODUCT = 4
UZEX_COL_QTY = 5
UZEX_COL_UNIT = 6
UZEX_COL_CONTRACT_TYPE = 7
UZEX_COL_STATUS = 8


def _decode_html_entities(text: str) -> str:
    return html.unescape(text)


def _parse_uzex_table(page_html: str) -> list[dict]:
    """Parse rows from UzEx HTML table. Returns list of dicts."""
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", page_html, re.DOTALL | re.IGNORECASE)
    records = []
    for row in rows:
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL | re.IGNORECASE)
        if len(cells) < 9:
            continue
        cells_text = [_decode_html_entities(re.sub(r"<[^>]+>", "", c)).strip() for c in cells]

        # Parse date: "17/04/2026 17:44:48"
        try:
            dt = datetime.strptime(cells_text[UZEX_COL_DATE][:16], "%d/%m/%Y %H:%M")
        except ValueError:
            continue

        try:
            deal_no = int(cells_text[UZEX_COL_DEAL_NO])
        except ValueError:
            continue

        try:
            price = float(cells_text[UZEX_COL_PRICE].replace(",", "").replace(" ", ""))
        except ValueError:
            price = None

        try:
            contract_no = int(cells_text[UZEX_COL_CONTRACT_NO])
        except ValueError:
            contract_no = None

        try:
            qty = float(cells_text[UZEX_COL_QTY].replace(",", "").replace(" ", ""))
        except ValueError:
            qty = None

        unit = cells_text[UZEX_COL_UNIT]

        records.append({
            "date": dt,
            "deal_no": deal_no,
            "price_sum": price,
            "contract_no": contract_no,
            "product": cells_text[UZEX_COL_PRODUCT],
            "qty": qty,
            "unit": unit,
            "contract_type": cells_text[UZEX_COL_CONTRACT_TYPE],
            "status": cells_text[UZEX_COL_STATUS],
        })
    return records


def fetch_uzex(search_term: str, days: int = 14) -> list[dict]:
    """
    Fetch UzEx records for given search term from the last N days.
    Paginates until records are older than cutoff date.
    """
    cutoff = datetime.now() - timedelta(days=days)
    all_records = []
    page = 1
    page_size = 1000

    print(f"  UzEx '{search_term}': fetching records from last {days} days...")

    while True:
        offset = (page - 1) * page_size
        url = (
            f"{UZEX_BASE}?Page={page}&Offset={offset}&Length={page_size}"
            f"&Search={urllib.parse.quote(search_term)}"
        )
        try:
            page_html = fetch(url)
        except Exception as e:
            print(f"    Ошибка при загрузке страницы {page}: {e}")
            break

        records = _parse_uzex_table(page_html)
        if not records:
            break

        new_in_range = [r for r in records if r["date"] >= cutoff]
        all_records.extend(new_in_range)

        oldest = min(r["date"] for r in records)
        print(f"    Страница {page}: {len(records)} записей, самая старая: {oldest.date()}, в диапазоне: {len(new_in_range)}")

        if oldest < cutoff:
            break
        page += 1

    print(f"  Итого записей в диапазоне: {len(all_records)}")
    return all_records


def update_uzex_excel(records: list[dict], sheet_name: str, dry_run: bool = False) -> int:
    """
    Update Excel sheet with UzEx records.
    - Adds new rows (by deal_no)
    - Updates status of existing rows from last 7 days
    Returns number of rows added/updated.
    """
    import xlwings as xw

    if not UZEX_FILE.exists():
        print(f"  Файл не найден: {UZEX_FILE}")
        return 0

    if dry_run:
        print(f"  [dry-run] Не записываю в {sheet_name}")
        for r in records[:5]:
            price_per_ton = r["price_sum"] / r["qty"] if r["qty"] and r["price_sum"] else 0
            print(f"    {r['date'].date()} | {r['deal_no']} | {r['product'][:40]} | "
                  f"{r['status']} | {r['qty']} {r['unit']} | {price_per_ton:,.0f} сум/т")
        return 0

    app = xw.App(visible=False)
    try:
        wb = app.books.open(str(UZEX_FILE))
        ws = wb.sheets[sheet_name]

        # Read existing deal numbers + row index
        last_row = ws.range("A1").end("down").row
        existing_deals: dict[int, int] = {}  # deal_no -> row_number
        if last_row > 1:
            deal_col = ws.range(f"B2:B{last_row}").value
            if not isinstance(deal_col, list):
                deal_col = [deal_col]
            for i, v in enumerate(deal_col):
                if v is not None:
                    try:
                        existing_deals[int(v)] = i + 2
                    except (ValueError, TypeError):
                        pass

        status_col = 9  # Column I = "Состояние исполнения" (1-based)
        update_cutoff = datetime.now() - timedelta(days=7)

        added = 0
        updated = 0

        for r in records:
            deal_no = r["deal_no"]
            if deal_no in existing_deals:
                # Update status if record is recent
                if r["date"] >= update_cutoff:
                    row_num = existing_deals[deal_no]
                    current_status = ws.range(f"I{row_num}").value
                    if current_status != r["status"]:
                        ws.range(f"I{row_num}").value = r["status"]
                        updated += 1
            else:
                # Append new row (columns A-I only; J and K are calculated by Excel formulas)
                next_row = last_row + 1
                ws.range(f"A{next_row}").value = r["date"]
                ws.range(f"B{next_row}").value = r["deal_no"]
                ws.range(f"C{next_row}").value = r["price_sum"]
                ws.range(f"D{next_row}").value = r["contract_no"]
                ws.range(f"E{next_row}").value = r["product"]
                ws.range(f"F{next_row}").value = r["qty"]
                ws.range(f"G{next_row}").value = r["unit"]
                ws.range(f"H{next_row}").value = r["contract_type"]
                ws.range(f"I{next_row}").value = r["status"]
                existing_deals[deal_no] = next_row
                last_row = next_row
                added += 1

        wb.save()
        wb.close()
        print(f"  {sheet_name}: добавлено {added}, обновлено статусов {updated}")
        return added + updated

    finally:
        app.quit()


# ============================================================
# ETS scraper
# ============================================================

ETS_URL = "https://ets.kz/markets/oil_products/trades/"

# ETS Excel columns in КАЗ_ETS_База sheet (1-based):
# A=Дата, B=Производитель, C=Вид транспорта, D=Марка битума,
# E=Мин цена (тенге), F=Макс цена (тенге), G=Ср цена (тенге),
# H=Объём (тонн), I=Кол-во сделок, J=Нач недели, K=Пятница, L=Год, M=№ недели


def _parse_ets_table(page_html: str) -> list[dict]:
    """
    Parse ETS trades table from HTML.
    ETS loads data dynamically via Bitrix AJAX.
    This function handles the static HTML table if present.
    """
    tables = re.findall(r"<table[^>]*>(.*?)</table>", page_html, re.DOTALL | re.IGNORECASE)
    records = []

    for table_html in tables:
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", table_html, re.DOTALL | re.IGNORECASE)
        for row in rows[1:]:  # Skip header row
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL | re.IGNORECASE)
            if len(cells) < 6:
                continue
            cells_text = [_decode_html_entities(re.sub(r"<[^>]+>", "", c)).strip() for c in cells]

            # Filter: only bitumen rows (product name contains битум/БНД/SBS)
            product = cells_text[3] if len(cells_text) > 3 else ""
            if not any(kw in product.lower() for kw in ("битум", "бнд", "sbs", "пмб")):
                continue

            try:
                dt = datetime.strptime(cells_text[0], "%d.%m.%Y")
            except ValueError:
                try:
                    dt = datetime.strptime(cells_text[0][:10], "%Y-%m-%d")
                except ValueError:
                    continue

            def safe_float(s):
                try:
                    return float(s.replace(" ", "").replace(",", "."))
                except ValueError:
                    return None

            week_start = dt - timedelta(days=dt.weekday())
            week_end = week_start + timedelta(days=4)
            week_no = dt.isocalendar()[1]

            records.append({
                "date": dt,
                "producer": cells_text[1] if len(cells_text) > 1 else "",
                "transport": cells_text[2] if len(cells_text) > 2 else "",
                "grade": product,
                "price_min": safe_float(cells_text[4]) if len(cells_text) > 4 else None,
                "price_max": safe_float(cells_text[5]) if len(cells_text) > 5 else None,
                "price_avg": safe_float(cells_text[6]) if len(cells_text) > 6 else None,
                "volume": safe_float(cells_text[7]) if len(cells_text) > 7 else None,
                "deals": safe_float(cells_text[8]) if len(cells_text) > 8 else None,
                "week_start": week_start,
                "week_end": week_end,
                "year": dt.year,
                "week_no": week_no,
            })

    return records


def _fetch_ets_with_playwright(date_from: str, date_to: str, timeout_ms: int = 30000) -> str:
    """
    Fetch ETS trades page via Playwright (headless Chromium).
    ETS loads trade data via JavaScript -- static HTTP scraping doesn't work.
    Returns page HTML after JS rendering, or empty string on error.
    """
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

    url = f"{ETS_URL}?date_from={date_from}&date_to={date_to}"
    print(f"  ETS Playwright: {url}")

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, timeout=timeout_ms)
            # Wait for the trade table to appear
            try:
                page.wait_for_selector(".i_markets_stock_table", timeout=15000)
            except PlaywrightTimeout:
                pass  # May not render -- still try to get HTML
            html = page.content()
            browser.close()
            return html
    except Exception as e:
        print(f"  ETS Playwright ошибка: {e}")
        return ""


def fetch_ets(days: int = 14) -> list[dict]:
    """
    Fetch ETS bitumen trade data via Playwright (headless browser).
    ETS uses JavaScript-rendered trade tables -- static HTTP scraping doesn't work.
    Returns list of trade records or empty list if unavailable.
    """
    cutoff = datetime.now() - timedelta(days=days)
    date_from = cutoff.strftime("%d.%m.%Y")
    date_to = datetime.now().strftime("%d.%m.%Y")

    print(f"  ETS: загружаю данные {date_from} -- {date_to}...")

    page_html = _fetch_ets_with_playwright(date_from, date_to)
    if not page_html:
        print("  ETS: не удалось получить данные.")
        return []

    records = _parse_ets_table(page_html)

    if not records:
        print("  ETS: строки с битумом не найдены в таблице.")
        print("  Проверьте: ets.kz/markets/oil_products/trades/ -- доступны ли данные за период?")
        return []

    print(f"  ETS: найдено {len(records)} записей с битумом за {days} дней")
    return records


def update_ets_excel(records: list[dict], dry_run: bool = False) -> int:
    """
    Update ETS_биржа.xlsx sheet КАЗ_ETS_База with new records.
    Appends only new rows (identified by date + producer + grade combination).
    """
    import xlwings as xw

    if not records:
        return 0

    if not ETS_FILE.exists():
        print(f"  Файл не найден: {ETS_FILE}")
        return 0

    if dry_run:
        print("  [dry-run] Не записываю в ETS_биржа.xlsx")
        for r in records[:5]:
            print(f"    {r['date'].date()} | {r['producer']} | {r['grade']} | {r['price_avg']} тг | {r['volume']} т")
        return 0

    app = xw.App(visible=False)
    try:
        wb = app.books.open(str(ETS_FILE))
        ws = wb.sheets["КАЗ_ETS_База"]

        last_row = ws.range("A1").end("down").row

        # Build set of existing (date_str, producer, grade) for dedup
        existing_keys: set[tuple] = set()
        if last_row > 1:
            for row_i in range(2, last_row + 1):
                d = ws.range(f"A{row_i}").value
                p = ws.range(f"B{row_i}").value or ""
                g = ws.range(f"D{row_i}").value or ""
                if d:
                    key = (str(d)[:10], str(p).strip(), str(g).strip())
                    existing_keys.add(key)

        added = 0
        for r in records:
            key = (str(r["date"])[:10], r["producer"].strip(), r["grade"].strip())
            if key in existing_keys:
                continue

            next_row = last_row + 1
            ws.range(f"A{next_row}").value = r["date"]
            ws.range(f"B{next_row}").value = r["producer"]
            ws.range(f"C{next_row}").value = r["transport"]
            ws.range(f"D{next_row}").value = r["grade"]
            ws.range(f"E{next_row}").value = r["price_min"]
            ws.range(f"F{next_row}").value = r["price_max"]
            ws.range(f"G{next_row}").value = r["price_avg"]
            ws.range(f"H{next_row}").value = r["volume"]
            ws.range(f"I{next_row}").value = r["deals"]
            ws.range(f"J{next_row}").value = r["week_start"]
            ws.range(f"K{next_row}").value = r["week_end"]
            ws.range(f"L{next_row}").value = r["year"]
            ws.range(f"M{next_row}").value = r["week_no"]
            existing_keys.add(key)
            last_row = next_row
            added += 1

        wb.save()
        wb.close()
        print(f"  КАЗ_ETS_База: добавлено {added} строк")
        return added

    finally:
        app.quit()


# ============================================================
# CCX scraper (Центральноазиатская товарно-сырьевая биржа, Казахстан)
# ============================================================

CCX_API = "https://ccx.kz/api/internal/trading-results"
CCX_SHEET = "КАЗ_CCX_База"

# CCX response fields → Excel columns (A-M):
# A=Дата, B=Код лота, C=Марка, D=Производитель, E=Базис поставки,
# F=Кол-во сделок, G=Мин цена (тг), H=Макс цена (тг), I=Ср цена сделок (тг),
# J=Объём сделок (тонн), K=Объём заявок продажа (тонн), L=Объём заявок покупка (тонн),
# M=Сумма сделок (тг)

CCX_HEADERS_ROW = [
    "Дата", "Код лота", "Марка", "Производитель", "Базис поставки",
    "Кол-во сделок", "Мин цена (тг)", "Макс цена (тг)", "Ср цена сделок (тг)",
    "Объём сделок (тонн)", "Объём заявок продажа (тонн)", "Объём заявок покупка (тонн)",
    "Сумма сделок (тг)",
]


def _parse_ccx_number(s: str) -> float | None:
    """Parse CCX number string: '161 445.78' or '2 015 тонна' → float."""
    if not s:
        return None
    # Remove units and spaces used as thousand separators
    cleaned = re.sub(r"\s*(тонна|тг|KZT).*$", "", str(s), flags=re.IGNORECASE)
    cleaned = cleaned.replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


CCX_COOKIE_FILE = BASE / ".ccx_cookies.json"


def _ccx_get_cookies_via_playwright() -> dict:
    """
    Open ccx.kz in headless Playwright, get session cookies, save to file.
    Called once to initialize cookies; subsequent calls reuse the saved file.
    """
    from playwright.sync_api import sync_playwright

    print("  CCX: получаю куки через Playwright (открываю ccx.kz)...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/147.0.0.0 Safari/537.36",
            locale="ru-RU",
        )
        page = ctx.new_page()
        try:
            page.goto("https://ccx.kz/gsm", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass
        cookies = {c["name"]: c["value"] for c in ctx.cookies()}
        browser.close()

    if cookies:
        CCX_COOKIE_FILE.write_text(json.dumps(cookies, ensure_ascii=False), encoding="utf-8")
        print(f"  CCX: куки сохранены ({len(cookies)} шт.)")
    return cookies


def _ccx_cookies() -> dict:
    """
    Load ccx.kz cookies: from saved file if fresh (<6h), else via Playwright.
    """
    if CCX_COOKIE_FILE.exists():
        age_hours = (datetime.now().timestamp() - CCX_COOKIE_FILE.stat().st_mtime) / 3600
        if age_hours < 6:
            return json.loads(CCX_COOKIE_FILE.read_text(encoding="utf-8"))
    return _ccx_get_cookies_via_playwright()


def _ccx_api_call(cookies: dict, max_results: int) -> list:
    """Make the CCX API request, return raw data list or empty list."""
    params = urllib.parse.urlencode({
        "section_code": "OIL",
        "instrument_regulated": "true",
        "max_results": max_results,
    })
    url = f"{CCX_API}?{params}"
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
        "Accept-Language": "ru-RU,ru;q=0.9",
        "Referer": "https://ccx.kz/gsm",
        "Cookie": "; ".join(f"{k}={v}" for k, v in cookies.items()),
    })
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=20) as r:
        data = json.loads(r.read().decode("utf-8"))
    if not data.get("success"):
        raise RuntimeError(f"API error: {data.get('message', '?')}")
    return data.get("data", [])


def _ccx_min_max(item: dict, mode: str) -> float | None:
    """
    Compute true min or max from all 4 price fields:
    Покупка (мин/макс) = bid, Продажа (мин/макс) = ask.
    """
    keys = (f"min_ask_price", f"min_bid_price") if mode == "min" else ("max_ask_price", "max_bid_price")
    vals = [_parse_ccx_number(item.get(k)) for k in keys]
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return min(vals) if mode == "min" else max(vals)


def fetch_ccx(max_results: int = 100) -> list[dict]:
    """
    Fetch CCX trading results via internal JSON API.
    Uses Playwright to get session cookies if not cached.
    Returns list of bitumen records.
    """
    print(f"  CCX: получаю данные (max_results={max_results})...")

    cookies = _ccx_cookies()
    if not cookies:
        print("  CCX: не удалось получить куки.")
        return []

    try:
        raw = _ccx_api_call(cookies, max_results)
    except Exception as e:
        print(f"  CCX ошибка (куки устарели?): {e}")
        # Try refreshing cookies once
        print("  CCX: обновляю куки...")
        CCX_COOKIE_FILE.unlink(missing_ok=True)
        cookies = _ccx_get_cookies_via_playwright()
        try:
            raw = _ccx_api_call(cookies, max_results)
        except Exception as e2:
            print(f"  CCX: повторная ошибка: {e2}")
            return []

    records = []
    for item in raw:
        lot_name = item.get("lot_name", "")
        if not any(kw in lot_name.lower() for kw in ("битум", "бнд", "sbs", "пмб")):
            continue
        try:
            dt = datetime.strptime(item["day_of_trade"], "%Y-%m-%d")
        except (ValueError, KeyError):
            continue
        records.append({
            "date": dt,
            "lot_code": item.get("lot_code", ""),
            "lot_name": lot_name,
            "basis": item.get("basis", ""),
            "delivery_term": item.get("delivery_term", ""),
            "deal_count": _parse_ccx_number(item.get("deal_count")),
            "min_price": _ccx_min_max(item, "min"),
            "max_price": _ccx_min_max(item, "max"),
            "avg_price": _parse_ccx_number(item.get("avg_price_deals")),
            "volume_deals": _parse_ccx_number(item.get("volume_deals")),
            "volume_ask": _parse_ccx_number(item.get("sum_ask_volume")),
            "volume_bid": _parse_ccx_number(item.get("sum_bid_volume")),
            "sum_deals": _parse_ccx_number(item.get("sum_deals")),
        })

    print(f"  CCX: найдено {len(records)} записей с битумом")
    return records


def _clean_ccx_producer(basis: str) -> str:
    """Normalize CCX basis/producer string to match ETS format."""
    # 'ТОО "СП" CASPI BITUM"' → 'Caspi Bitum'
    s = basis.upper()
    if "CASPI" in s or "КАСПИ" in s:
        return "Caspi Bitum"
    # Strip legal form prefixes (ТОО, АО, ЗАО)
    s = re.sub(r'^(ТОО|АО|ЗАО|ООО)\s*["\']?', "", basis.strip(), flags=re.IGNORECASE).strip(' "\'')
    return s or basis.strip()


def update_ccx_excel(records: list[dict], dry_run: bool = False) -> int:
    """
    Append CCX records to unified КАЗ_ETS_База sheet in ETS_биржа.xlsx.
    Maps CCX fields to ETS column layout:
      A=Дата, B=Производитель, C=Вид транспорта, D=Марка битума,
      E=Мин цена тенге, F=Макс цена тенге, G=Ср цена тенге,
      H=Объём тонн, I=Кол-во сделок, J=Нач недели, K=Пятница, L=Год, M=№ недели
    Uses openpyxl (ETS_биржа.xlsx has no Power Query).
    """
    import openpyxl

    if not records:
        return 0

    if not ETS_FILE.exists():
        print(f"  Файл не найден: {ETS_FILE}")
        return 0

    if dry_run:
        print(f"  [dry-run] Не записываю в КАЗ_ETS_База")
        for r in records[:5]:
            producer = _clean_ccx_producer(r["basis"])
            mn = f"{r['min_price']:,.0f}" if r.get("min_price") else "?"
            mx = f"{r['max_price']:,.0f}" if r.get("max_price") else "?"
            avg = f"{r['avg_price']:,.0f}" if r.get("avg_price") else "?"
            print(f"    {r['date'].date()} | {producer} | мин={mn} макс={mx} ср={avg} тг | {r['volume_deals']} т")
        return 0

    wb = openpyxl.load_workbook(str(ETS_FILE))
    ws = wb["КАЗ_ETS_База"]

    # Build dedup set: (date_str, producer, grade)
    existing_keys: set[tuple] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d, p, _, g = row[0], row[1], row[2], row[3]
        if d:
            existing_keys.add((str(d)[:10], str(p or "").strip(), str(g or "").strip()))

    # Remove erroneous CCX sheet if it exists
    if CCX_SHEET in wb.sheetnames:
        del wb[CCX_SHEET]
        print(f"  Удалён лишний лист {CCX_SHEET}")

    added = 0
    for r in records:
        producer = _clean_ccx_producer(r["basis"])
        grade = r["lot_name"]
        key = (r["date"].strftime("%Y-%m-%d"), producer, grade)
        if key in existing_keys:
            continue

        dt = r["date"]
        week_start = dt - timedelta(days=dt.weekday())      # Monday
        week_end = week_start + timedelta(days=4)            # Friday
        week_no = dt.isocalendar()[1]

        ws.append([
            dt,           # A Дата
            producer,     # B Производитель
            "CCX",        # C Вид транспорта (источник)
            grade,        # D Марка битума
            r["min_price"],   # E Мин цена тенге
            r["max_price"],   # F Макс цена тенге
            r["avg_price"],   # G Ср цена тенге
            r["volume_deals"],# H Объём тонн
            r["deal_count"],  # I Кол-во сделок
            week_start,   # J Нач недели
            week_end,     # K Пятница
            dt.year,      # L Год
            week_no,      # M № недели
        ])
        existing_keys.add(key)
        added += 1

    wb.save(str(ETS_FILE))
    print(f"  КАЗ_ETS_База: добавлено {added} строк из CCX")
    return added


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Exchange scraper: UzEx + ETS + CCX")
    parser.add_argument("--uzex", action="store_true", help="Only UzEx")
    parser.add_argument("--ets", action="store_true", help="Only ETS")
    parser.add_argument("--ccx", action="store_true", help="Only CCX")
    parser.add_argument("--days", type=int, default=14, help="Days lookback for UzEx (default 14)")
    parser.add_argument("--max-results", type=int, default=100, help="Max records from CCX (default 100, API limit)")
    parser.add_argument("--dry-run", action="store_true", help="Show data without writing to Excel")
    args = parser.parse_args()

    run_all = not args.uzex and not args.ets and not args.ccx
    run_uzex = args.uzex or run_all
    run_ets = args.ets or run_all
    run_ccx = args.ccx or run_all

    total_changed = 0

    # --- UzEx ---
    if run_uzex:
        print("\n[UzEx -- Узбекская товарно-сырьевая биржа]")
        bitum_records = fetch_uzex("битум", days=args.days)
        if bitum_records:
            total_changed += update_uzex_excel(bitum_records, "Битум", dry_run=args.dry_run)
        mazut_records = fetch_uzex("мазут", days=args.days)
        if mazut_records:
            total_changed += update_uzex_excel(mazut_records, "Мазут", dry_run=args.dry_run)

    # --- ETS ---
    if run_ets:
        print("\n[ETS -- Казахстанская энергетическая биржа]")
        ets_records = fetch_ets(days=args.days)
        if ets_records:
            total_changed += update_ets_excel(ets_records, dry_run=args.dry_run)

    # --- CCX ---
    if run_ccx:
        print("\n[CCX -- Центральноазиатская товарно-сырьевая биржа]")
        ccx_records = fetch_ccx(max_results=args.max_results)
        if ccx_records:
            total_changed += update_ccx_excel(ccx_records, dry_run=args.dry_run)

    print(f"\nГотово. Всего изменений: {total_changed}")


if __name__ == "__main__":
    main()
